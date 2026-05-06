import os
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

base_dir = "/Users/vishalmahto/Desktop/AIML_starterKit/paid_Bundle/AI-Engineer-Foundation-System"

# Actual links for resources
LINKS = {
    "English": [
        "https://www.freecodecamp.org/news/tag/machine-learning/",
        "https://www.coursera.org/specializations/machine-learning-introduction",
        "https://neetcode.io/"
    ],
    "Hindi": [
        "https://www.youtube.com/c/CodeWithHarry",
        "https://www.youtube.com/user/krishnaik06"
    ]
}

def generate_markdown_content(filename):
    title = filename.replace(".pdf", "").replace("_", " ")
    
    # Embedding CSS for larger font size and styling
    content = f"""<style>
    body {{
        font-size: 18px;
        line-height: 1.8;
        font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
        color: #1f2937;
        margin: 40px;
    }}
    h1 {{ font-size: 34px; color: #1d4ed8; border-bottom: 2px solid #e5e7eb; padding-bottom: 12px; margin-bottom: 20px; }}
    h2 {{ font-size: 26px; color: #1e40af; margin-top: 35px; margin-bottom: 15px; }}
    h3 {{ font-size: 22px; color: #1e3a8a; margin-top: 25px; margin-bottom: 10px; }}
    ul, ol {{ margin-left: 25px; margin-bottom: 20px; }}
    li {{ margin-bottom: 12px; }}
    p {{ margin-bottom: 18px; }}
    a {{ color: #2563eb; text-decoration: none; font-weight: 600; border-bottom: 1px solid transparent; transition: border-color 0.2s; }}
    a:hover {{ border-bottom-color: #2563eb; }}
    hr {{ border: 0; height: 1px; background: #e5e7eb; margin: 40px 0; }}
    .footer {{ font-size: 14px; color: #6b7280; font-style: italic; text-align: center; margin-top: 60px; }}
    </style>

# {title}

Welcome to the **{title}** module of the AI Engineer Foundation System. 
This document provides everything you need to master this topic, moving you closer to becoming a full-stack AI Engineer.

---

## 🎯 Learning Objectives
* Understand the core concepts behind {title}.
* Apply these concepts practically in real-world AI/ML contexts.
* Avoid common pitfalls that trap beginners.

## 📚 Curriculum & Study Plan

### 1. Conceptual Foundation
You must understand the theory before writing the code. 
* **Focus Area:** Grasp the underlying mechanics.
* **Why it matters:** AI requires mathematical and logical intuition, not just syntax memory.

### 2. Required Resources
We have curated the best resources so you don't waste time in tutorial hell. Click the links below to access the study material directly:

**English Resources:**
* [**freeCodeCamp Machine Learning Hub**]({LINKS['English'][0]}) - Comprehensive text and video tutorials.
* [**Andrew Ng Machine Learning Specialization**]({LINKS['English'][1]}) - For theoretical depth and mathematical foundations.
* [**NeetCode (DSA)**]({LINKS['English'][2]}) - For parallel Data Structures and Algorithms practice.

**Hindi Resources:**
* [**CodeWithHarry YouTube Channel**]({LINKS['Hindi'][0]}) - Excellent beginner-friendly programming content.
* [**Krish Naik YouTube Channel**]({LINKS['Hindi'][1]}) - Detailed AI/ML concepts and practical implementation walkthroughs.

### 3. Execution & Practice
Do not just watch videos. You must build.
* **Task:** Implement the concepts from `{title}` in a blank Python script or Jupyter Notebook.
* **Rule:** If you get stuck, use the 15-minute rule before checking the solution.

## 🛠 Project / Implementation
To solidify this knowledge, you will build a mini-project.
1. **Define the Problem:** How does `{title}` solve a real problem?
2. **Write the Code:** Implement it from scratch.
3. **Refactor:** Clean up your code, add docstrings, and push to GitHub.

---
<div class="footer">
AI Engineer Foundation System - Premium Digital Product<br>
Designed for execution, discipline, and portfolio building.
</div>
"""
    return content

def update_xlsx_to_do(filepath, title):
    wb = Workbook()
    ws = wb.active
    ws.title = "To-Do List"
    
    # Headers
    headers = ["Task ID", "Topic/Concept", "Resource Link", "Status", "Priority", "Deadline", "Notes"]
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    # Sample Rows
    sample_tasks = [
        [1, f"Review theory for {title}", "https://www.freecodecamp.org/", "Not Started", "High", "Day 1", "Focus on core concepts"],
        [2, f"Watch Hindi tutorial for {title}", "https://www.youtube.com/c/CodeWithHarry", "Not Started", "Medium", "Day 1", "Take handwritten notes"],
        [3, "Implement 3 mini-exercises", "Local Code Editor", "Not Started", "High", "Day 2", "Apply 15-min stuck rule"],
        [4, "Complete project implementation", "GitHub Repository", "Not Started", "High", "Day 3", "Push clean, documented code"]
    ]
    
    for row in sample_tasks:
        ws.append(row)
        
    # Adjust column widths
    column_widths = {'A': 10, 'B': 35, 'C': 45, 'D': 15, 'E': 15, 'F': 15, 'G': 40}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
        
    wb.save(filepath)

# Iterate through all files and update
for root, dirs, files in os.walk(base_dir):
    for file in files:
        filepath = os.path.join(root, file)
        
        if file.endswith('.pdf'):
            md_path = filepath.replace('.pdf', '.md')
            with open(md_path, 'w') as f:
                f.write(generate_markdown_content(file))
            
            print(f"Updating PDF for {file}...")
            try:
                subprocess.run(["npx", "md-to-pdf", md_path], cwd=root, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            except Exception as e:
                print(f"Failed to generate {file}: {e}")
            
            if os.path.exists(md_path):
                os.remove(md_path)
                
        elif file.endswith('.xlsx'):
            print(f"Updating XLSX for {file}...")
            title = file.replace(".xlsx", "").replace("_", " ")
            try:
                update_xlsx_to_do(filepath, title)
            except Exception as e:
                print(f"Failed to update {file}: {e}")

print("All PDFs updated with larger fonts and links. All XLSX files updated with To-Do lists.")
